"""
Internal Training Survey Form — ABC-MIB Group Co., Ltd.

Streamlit + Supabase (Auth + Postgres) production app.
Run: streamlit run app.py
"""

import base64
import io
import time
from datetime import date, datetime, timedelta
from typing import Optional
from zoneinfo import ZoneInfo

import httpx
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import Client, ClientOptions, create_client

# The Supabase auth (gotrue) client builds its own httpx client with a
# short, fixed default timeout unless one is passed in explicitly -- this
# is unaffected by postgrest_client_timeout and easy to exceed on
# cross-region latency between Streamlit Cloud and a Supabase project,
# surfacing as "the read operation timed out" on sign-up/sign-in. Passing
# a shared httpx client via ClientOptions applies this same timeout to
# every sub-client (auth, postgrest, storage) created from it.
SUPABASE_HTTP_TIMEOUT_SECONDS = 30

# Submissions are stored/fetched in UTC; the dashboard's Date Range filter
# compares against calendar dates in this timezone so "today" means the
# same thing to the admin using it as it did to the trainee submitting.
LOCAL_TZ = ZoneInfo("Asia/Yangon")

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="ABC-MIB Training Survey",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)

COMPANY_NAME = st.secrets.get("app", {}).get("company_name", "ABC-MIB Group Co., Ltd.")

# Deep Emerald & Slate brand palette. Also mirrored in .streamlit/config.toml
# (theme primaryColor etc.) for native widget theming -- these hex values
# here are for the custom CSS below and for coloring the analytics bar
# charts. Both EMERALD and SLATE independently pass the dataviz skill's
# standalone contrast check (>=3:1 against a white chart surface), so
# unlike the previous gold-based palette, no separate darker "chart-safe"
# variant is needed here.
EMERALD = "#0B4F3C"
EMERALD_DARK = "#073A2C"
SLATE = "#3A4750"
SLATE_LIGHT = "#C7D0D6"
SURFACE = "#F5F7F6"
BORDER = "#E1E6E3"
TEXT_MUTED = "#5B6B66"

RATING_OPTIONS = ["4 - Excellent", "3 - Good", "2 - Average", "1 - Poor"]
RATING_QUESTIONS = [
    ("q1_enjoyable", "1. I enjoyed the training session."),
    ("q2_organized", "2. The training was well organized."),
    ("q3_satisfied", "3. Overall, I am satisfied with this training."),
    ("q4_improved_understanding", "4. This training improved my understanding of the subject."),
    ("q5_respect_ideas", "5. My ideas and opinions were respected during the session."),
    ("q6_presenter_rating", "6. How would you rate the presenter?"),
]
OPEN_QUESTIONS = [
    ("q7_understanding_notes", "7. What did you understand from this training?"),
    ("q8_interesting_thing", "8. What was the most interesting thing you learned?"),
    ("q9_knowledge_application", "9. How will you apply this knowledge in your work?"),
    ("q10_challenge_addressed", "10. What challenge (if any) did this training help address?"),
    ("q11_future_topics", "11. What topics would you like to see in future training?"),
]

CUSTOM_CSS = f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    html, body, [class*="css"] {{
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    }}

    .block-container {{padding-top: 2rem; padding-bottom: 3rem;}}

    /* Narrow, centered "card" shell for trainee-facing pages (login, survey
       form). The app uses wide layout so the admin dashboard's tables and
       charts get full desktop width; this container opts specific pages
       back into a constrained, polished width instead of a blanket
       .block-container rule that would cramp the admin views too. */
    .st-key-app-shell {{
        max-width: 720px;
        margin: 0 auto;
    }}

    .survey-header {{
        text-align: center;
        background: linear-gradient(135deg, {EMERALD} 0%, {EMERALD_DARK} 100%);
        color: #FFFFFF;
        padding: 1.75rem 1.5rem 1.5rem;
        border-radius: 14px;
        margin-bottom: 1.75rem;
        box-shadow: 0 4px 16px rgba(15, 42, 74, 0.18);
    }}
    .survey-header h1 {{
        font-size: 1.3rem;
        font-weight: 700;
        margin: 0 0 0.2rem;
        letter-spacing: 0.02em;
    }}
    .survey-header h3 {{
        font-size: 1rem;
        font-weight: 400;
        margin: 0;
        color: {SLATE_LIGHT};
    }}
    .survey-header::after {{
        content: "";
        display: block;
        width: 56px;
        height: 3px;
        background: {SLATE_LIGHT};
        margin: 0.9rem auto 0;
        border-radius: 2px;
    }}

    /* Form card */
    div[data-testid="stForm"] {{
        border: 1px solid {BORDER};
        border-radius: 14px;
        padding: 1.75rem;
        background: #FFFFFF;
        box-shadow: 0 2px 10px rgba(15, 42, 74, 0.06);
    }}

    /* Sidebar */
    section[data-testid="stSidebar"] {{
        background: {EMERALD};
    }}
    section[data-testid="stSidebar"] * {{
        color: {SURFACE} !important;
    }}
    section[data-testid="stSidebar"] button {{
        border-color: {SLATE_LIGHT} !important;
    }}

    /* Tabs: slate underline on the selected tab, emerald label */
    button[data-baseweb="tab"][aria-selected="true"] {{
        color: {EMERALD} !important;
        font-weight: 600;
    }}
    div[data-baseweb="tab-highlight"] {{
        background-color: {SLATE} !important;
        height: 3px !important;
    }}

    /* Metric tiles: card look for dashboard Overview rows */
    div[data-testid="stMetric"] {{
        background: #FFFFFF;
        border: 1px solid {BORDER};
        border-radius: 12px;
        padding: 0.9rem 1rem;
        box-shadow: 0 2px 8px rgba(15, 42, 74, 0.05);
    }}
    div[data-testid="stMetricLabel"] {{
        color: {TEXT_MUTED};
    }}

    /* Tables */
    div[data-testid="stDataFrame"] {{
        border: 1px solid {BORDER};
        border-radius: 10px;
        overflow: hidden;
    }}

    /* Expander rows (My Submissions / admin's View Submissions) */
    details[data-testid="stExpander"] {{
        border: 1px solid {BORDER} !important;
        border-radius: 10px !important;
    }}

    /* Disabled text inputs (locked Name/Position/Department): Streamlit/
       Baseweb washes these out by default -- light gray text, and on
       WebKit/Blink specifically via -webkit-text-fill-color rather than
       just `color`, so overriding only `color` doesn't fix it -- which
       reads as barely-visible against our light SURFACE input background.
       Force legible dark emerald text; a slate left border communicates
       "read-only" instead of relying on faded text to do it. */
    div[data-testid="stTextInput"] input:disabled {{
        color: {EMERALD} !important;
        -webkit-text-fill-color: {EMERALD} !important;
        opacity: 1 !important;
        background-color: {SURFACE} !important;
        border-left: 3px solid {SLATE} !important;
    }}
    div[data-testid="stTextInput"] label {{
        color: {EMERALD} !important;
    }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Supabase client
# ---------------------------------------------------------------------------
# IMPORTANT: do NOT wrap this in @st.cache_resource. Streamlit Community
# Cloud runs one Python process shared by every visitor; a cached client
# would carry a single shared auth session and leak one user's login into
# another user's browser tab. Each browser session gets its own client,
# stored in that session's st.session_state.
def _is_privileged_key(key: str) -> bool:
    """Detect a secret/service_role key, which bypasses RLS and must never
    be used in a public-facing app."""
    if key.startswith("sb_secret_"):
        return True
    parts = key.split(".")
    if len(parts) == 3:  # legacy JWT-style key (anon / service_role)
        try:
            padded = parts[1] + "=" * (-len(parts[1]) % 4)
            payload = base64.urlsafe_b64decode(padded).decode("utf-8", "ignore")
            if '"role":"service_role"' in payload:
                return True
        except Exception:
            pass
    return False


def get_supabase() -> Client:
    if "sb_client" not in st.session_state:
        supabase_secrets = st.secrets.get("supabase", {})
        url = supabase_secrets.get("url")
        key = supabase_secrets.get("key")

        if not url or not key or "YOUR_SUPABASE" in url or "YOUR_SUPABASE" in key:
            st.error(
                "**Supabase credentials are not configured.**\n\n"
                "- On Streamlit Community Cloud: open **Manage app → Settings → Secrets** "
                "and paste your `[supabase]` url and key.\n"
                "- Running locally: fill in the real values in `.streamlit/secrets.toml`.\n\n"
                "See README.md section 4/6 for the exact format."
            )
            st.stop()

        if _is_privileged_key(key):
            st.error(
                "**The configured Supabase key is a secret/service_role key.** "
                "It bypasses Row Level Security, so using it here would let any "
                "visitor read or modify all data — never use it in a public app.\n\n"
                "In Supabase: **Project Settings → API Keys**, copy the "
                "**Publishable key** (`sb_publishable_...`) — or the legacy "
                "**anon public** key — and use that instead."
            )
            st.stop()

        options = ClientOptions(httpx_client=httpx.Client(timeout=SUPABASE_HTTP_TIMEOUT_SECONDS))
        st.session_state.sb_client = create_client(url, key, options=options)
    return st.session_state.sb_client


def delete_user(user_id: str) -> None:
    """Permanently deletes a user via the Supabase Auth Admin API.

    This is the one deliberate, narrow use of the secret/service_role key
    in the whole app: a fresh admin client is built here, used for exactly
    this one call, and discarded — never stored in session_state, never
    used for anything else, and only ever reached from admin-gated UI.
    survey_responses keeps the person's past submissions (user_id is set
    to NULL by the FK's ON DELETE SET NULL), only the login account and
    users_profile row (cascades) are removed.
    """
    service_key = st.secrets.get("supabase", {}).get("service_role_key")
    if not service_key:
        raise RuntimeError(
            "No service_role key configured. Add service_role_key under "
            "[supabase] in secrets to enable deleting users."
        )
    url = st.secrets["supabase"]["url"]
    options = ClientOptions(httpx_client=httpx.Client(timeout=SUPABASE_HTTP_TIMEOUT_SECONDS))
    admin_client = create_client(url, service_key, options=options)
    admin_client.auth.admin.delete_user(user_id)


# ---------------------------------------------------------------------------
# Session / auth helpers
# ---------------------------------------------------------------------------
def _call_with_retry(func, attempts: int = 3, delay_seconds: float = 1.5):
    """Retries a flaky network call a couple of times before giving up.

    The Supabase auth client (gotrue) builds its own httpx client with a
    default 5-second timeout and no retry, unlike the Postgres/table client.
    On Streamlit Cloud, or whenever a free-tier Supabase project has been
    idle and needs to spin back up, that's often not enough time for the
    very first request and it fails with a plain read-timeout — even though
    a near-immediate retry usually succeeds. Only used for calls where
    retrying a request Supabase might have actually completed is safe
    (sign-in is idempotent; sign-up is handled separately, see sign_up()).
    """
    last_exc: Exception = RuntimeError("no attempt made")
    for attempt in range(attempts):
        try:
            return func()
        except Exception as e:
            last_exc = e
            if attempt < attempts - 1:
                time.sleep(delay_seconds)
    raise last_exc


def init_session_state():
    defaults = {
        "auth_user": None,      # {"id", "email"}
        "profile": None,        # row from users_profile
        "form_nonce": 0,        # bumped to reset the survey form widgets
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def load_profile(sb: Client, user_id: str):
    resp = sb.table("users_profile").select("*").eq("id", user_id).single().execute()
    return resp.data


def sign_in(sb: Client, email: str, password: str):
    # Signing in is safe to retry -- it doesn't create anything server-side,
    # so a request that timed out client-side but actually succeeded, or one
    # that simply needs a moment because the project was idle, both resolve
    # cleanly on a second attempt instead of showing a raw timeout error.
    result = _call_with_retry(
        lambda: sb.auth.sign_in_with_password({"email": email, "password": password})
    )
    st.session_state.auth_user = {"id": result.user.id, "email": result.user.email}
    st.session_state.profile = load_profile(sb, result.user.id)


def sign_up(sb: Client, email: str, password: str, full_name: str, department: str, position: str):
    # full_name/department/position are passed as Supabase Auth user
    # metadata so the handle_new_user() trigger in schema.sql can save
    # them even when email confirmation is enabled and there's no active
    # session yet to authorize a direct table write. That only takes
    # effect once schema.sql has actually been re-run against the
    # database, though -- so as a second, independent path that works
    # immediately regardless of that, we also write directly to
    # users_profile below whenever sign-up returns a session right away.
    signup_payload = {
        "email": email,
        "password": password,
        "options": {
            "data": {
                "full_name": full_name,
                "department": department,
                "position": position,
            }
        },
    }

    # Unlike sign-in, a sign-up call isn't safe to blindly retry in a loop:
    # if the first attempt actually reached Supabase and created the account
    # before the client's 5-second timeout gave up on it, retrying the exact
    # same call fails with "already registered" -- which we treat as
    # confirmation the account exists and tell the user to log in instead,
    # rather than surfacing a confusing raw timeout.
    try:
        result = sb.auth.sign_up(signup_payload)
    except Exception as first_error:
        try:
            result = sb.auth.sign_up(signup_payload)
        except Exception as second_error:
            if "already" in str(second_error).lower() or "already" in str(first_error).lower():
                raise RuntimeError(
                    "This email may already be registered — a previous attempt likely went "
                    "through despite the timeout. Try logging in instead."
                ) from second_error
            raise RuntimeError(
                "The server took too long to respond, twice in a row. This can happen if the "
                "database was waking up from being idle. Please wait a few seconds and try again."
            ) from second_error

    if result.user is None:
        raise RuntimeError("Sign-up did not return a user. Please try again.")

    if result.session is not None and result.user.id:
        try:
            sb.table("users_profile").update(
                {"full_name": full_name, "department": department, "position": position}
            ).eq("id", result.user.id).execute()
        except Exception:
            pass  # best-effort direct save; the trigger-based path may still have covered it
        st.session_state.auth_user = {"id": result.user.id, "email": result.user.email}
        st.session_state.profile = load_profile(sb, result.user.id)
        return True  # logged in immediately
    return False  # needs email confirmation


def sign_out(sb: Client):
    try:
        sb.auth.sign_out()
    except Exception:
        pass
    for k in ("auth_user", "profile", "sb_client"):
        st.session_state.pop(k, None)
    st.rerun()


# ---------------------------------------------------------------------------
# Auth screen
# ---------------------------------------------------------------------------
def render_auth_screen(sb: Client):
    st.markdown(
        f"""
        <div class="survey-header">
            <h1>{COMPANY_NAME}</h1>
            <h3>Internal Training Survey Form</h3>
        </div>
        """,
        unsafe_allow_html=True,
    )

    login_tab, signup_tab = st.tabs(["Log In", "Sign Up"])

    with login_tab:
        with st.form("login_form"):
            email = st.text_input("Email", key="login_email")
            password = st.text_input("Password", type="password", key="login_password")
            submitted = st.form_submit_button("Log In", use_container_width=True, type="primary")
        if submitted:
            if not email or not password:
                st.error("Please enter both email and password.")
            else:
                try:
                    sign_in(sb, email, password)
                    st.rerun()
                except Exception as e:
                    st.error(f"Login failed: {e}")

    with signup_tab:
        with st.form("signup_form"):
            full_name = st.text_input("Full Name", key="signup_name")
            col1, col2 = st.columns(2)
            with col1:
                department = st.text_input("Department", key="signup_dept")
            with col2:
                position = st.text_input("Position", key="signup_position")
            email = st.text_input("Email", key="signup_email")
            password = st.text_input("Password", type="password", key="signup_password")
            password2 = st.text_input("Confirm Password", type="password", key="signup_password2")
            submitted = st.form_submit_button("Create Account", use_container_width=True, type="primary")
        if submitted:
            if not all([full_name, email, password, password2]):
                st.error("Please fill in all required fields.")
            elif password != password2:
                st.error("Passwords do not match.")
            elif len(password) < 6:
                st.error("Password must be at least 6 characters.")
            else:
                try:
                    logged_in = sign_up(sb, email, password, full_name, department, position)
                    if logged_in:
                        st.rerun()
                    else:
                        st.success("Account created! Please check your email to confirm, then log in.")
                except Exception as e:
                    st.error(f"Sign-up failed: {e}")


# ---------------------------------------------------------------------------
# Survey form (trainee)
# ---------------------------------------------------------------------------
def render_survey_form(sb: Client):
    st.markdown(
        f"""
        <div class="survey-header">
            <h1>{COMPANY_NAME}</h1>
            <h3>Internal Training Survey Form</h3>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.write(f"Logged in as **{st.session_state.auth_user['email']}**")
        if st.button("🔄 Refresh", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
        if st.button("Log Out", use_container_width=True):
            sign_out(sb)

    tab_new, tab_mine = st.tabs(["📝 New Survey", "🗂️ My Submissions"])
    with tab_new:
        render_new_survey_tab(sb)
    with tab_mine:
        render_my_submissions_tab(sb)


def render_new_survey_tab(sb: Client):
    # A toast alone is easy to miss (small, corner-of-screen, fades fast).
    # Show a persistent banner too, exactly once, on the freshly-reset form
    # right after a successful submission.
    if st.session_state.pop("just_submitted", False):
        st.success("✅ Survey submitted successfully — thank you!")

    # Fetch fresh rather than trusting the cached session_state.profile, so
    # a fix an admin just made via Manage Users -> Edit Info shows up here
    # immediately instead of only after the trainee logs out and back in.
    try:
        profile = load_profile(sb, st.session_state.auth_user["id"]) or {}
        st.session_state.profile = profile
    except Exception:
        profile = st.session_state.profile or {}

    full_name = profile.get("full_name") or ""
    position_value = profile.get("position") or ""
    department_value = profile.get("department") or ""

    if not (full_name and position_value and department_value):
        st.warning(
            "Your profile is missing Name/Position/Department, so those fields below are blank. "
            "These are locked and can only be filled in by an admin — ask one to complete your "
            "profile (Manage Users → Edit Info) before you can submit."
        )

    nonce = st.session_state.form_nonce

    with st.form(f"survey_form_{nonce}", clear_on_submit=False):
        st.subheader("Basic Information")

        trainee_name = st.text_input(
            "Name *", value=full_name, key=f"name_{nonce}", disabled=True
        )
        c1, c2 = st.columns(2)
        with c1:
            position = st.text_input(
                "Position", value=position_value, key=f"pos_{nonce}", disabled=True
            )
        with c2:
            department = st.text_input(
                "Department", value=department_value, key=f"dept_{nonce}", disabled=True
            )

        training_course = st.text_input("Training Course *", key=f"course_{nonce}")
        presenter_name = st.text_input("Presenter Name", key=f"presenter_{nonce}")

        c3, c4 = st.columns(2)
        with c3:
            start_date = st.date_input("Start Date", value=date.today(), key=f"start_{nonce}")
        with c4:
            completed_date = st.date_input("Completed Date", value=date.today(), key=f"completed_{nonce}")

        c5, c6 = st.columns(2)
        with c5:
            training_hours = st.number_input(
                "Training Hours", min_value=0.0, step=0.5, key=f"hours_{nonce}"
            )
        with c6:
            time_range = st.text_input("Time Range (e.g. 09:00-12:00)", key=f"time_{nonce}")

        st.divider()
        st.subheader("Please rate the following")
        st.caption("4 = Excellent · 3 = Good · 2 = Average · 1 = Poor")

        rating_values = {}
        for field, label in RATING_QUESTIONS:
            st.markdown(f"**{label}** *")
            rating_values[field] = st.segmented_control(
                label, options=RATING_OPTIONS, key=f"{field}_{nonce}", label_visibility="collapsed"
            )

        st.divider()
        st.subheader("Your Feedback")

        open_values = {}
        for field, label in OPEN_QUESTIONS:
            open_values[field] = st.text_area(label, key=f"{field}_{nonce}", height=80)

        submitted = st.form_submit_button("Submit Survey", use_container_width=True, type="primary")

    if not submitted:
        return

    # ---- validation ----
    missing = []
    if not trainee_name.strip():
        missing.append("Trainee Name")
    if not training_course.strip():
        missing.append("Training Course")
    for field, label in RATING_QUESTIONS:
        if rating_values[field] is None:
            missing.append(label)

    if missing:
        st.error("Please complete the following required fields:\n\n- " + "\n- ".join(missing))
        return

    def rating_to_int(selected: str) -> int:
        return int(selected.split(" - ")[0])

    payload = {
        "user_id": st.session_state.auth_user["id"],
        "trainee_name": trainee_name.strip(),
        "position": position.strip(),
        "department": department.strip(),
        "training_course": training_course.strip(),
        "start_date": start_date.isoformat(),
        "completed_date": completed_date.isoformat(),
        "training_hours": training_hours,
        "time_range": time_range.strip(),
        "presenter_name": presenter_name.strip(),
        **{field: rating_to_int(rating_values[field]) for field, _ in RATING_QUESTIONS},
        **{field: open_values[field].strip() for field, _ in OPEN_QUESTIONS},
    }

    try:
        sb.table("survey_responses").insert(payload).execute()
        st.toast("Survey submitted — thank you!", icon="✅")
        st.session_state.just_submitted = True
        st.session_state.form_nonce += 1  # forces a clean form on rerun
        st.rerun()
    except Exception as e:
        st.error(f"Could not submit survey: {e}")


def fetch_submissions_for_user(sb: Client, user_id: str) -> list[dict]:
    resp = (
        sb.table("survey_responses")
        .select("*")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
        .execute()
    )
    return resp.data or []


def delete_submission(sb: Client, response_id: str) -> None:
    """Permanently deletes one survey response. Backed by the
    responses_delete_admin RLS policy in schema.sql, so this works with
    the regular publishable-key client — no privileged key needed."""
    sb.table("survey_responses").delete().eq("id", response_id).execute()


def render_delete_submission_control(sb: Client, response_id: str, key_prefix: str):
    """Two-step delete confirm for one submission, reusable wherever an
    admin is looking at individual responses."""
    confirm_key = f"{key_prefix}_confirm_{response_id}"

    if not st.session_state.get(confirm_key):
        if st.button("🗑️ Delete Submission", key=f"{key_prefix}_btn_{response_id}", use_container_width=True):
            st.session_state[confirm_key] = True
            st.rerun()
        return

    st.warning("Permanently delete this submission? This cannot be undone.")
    dc1, dc2 = st.columns(2)
    with dc1:
        confirmed = st.button(
            "Yes, delete", key=f"{key_prefix}_yes_{response_id}", use_container_width=True, type="primary"
        )
    with dc2:
        if st.button("Cancel", key=f"{key_prefix}_no_{response_id}", use_container_width=True):
            st.session_state[confirm_key] = False
            st.rerun()

    if confirmed:
        try:
            delete_submission(sb, response_id)
            st.cache_data.clear()
            st.session_state[confirm_key] = False
            st.toast("Submission deleted", icon="🗑️")
            st.rerun()
        except Exception as e:
            st.error(f"Could not delete submission ({type(e).__name__}): {e}")


def render_submission_list(rows: list[dict], sb: Optional[Client] = None, allow_delete: bool = False):
    """Shared expander-list layout for a set of submissions — used both for
    a trainee's own 'My Submissions' tab and an admin viewing someone
    else's submissions the same way, from Manage Users. allow_delete is
    admin-only: trainees never get a delete control on their own view."""
    if not rows:
        st.info("No submissions yet.")
        return

    st.caption(f"{len(rows)} submission(s)")

    for row in rows:
        # Raw dict from Supabase (not the cached fetch_responses DataFrame),
        # so it needs its own UTC -> local conversion before display.
        created = pd.to_datetime(row["created_at"]).tz_convert(LOCAL_TZ).strftime("%Y-%m-%d %H:%M")
        with st.expander(f"{created} — {row.get('training_course') or '(no course)'}"):
            c1, c2 = st.columns(2)
            with c1:
                st.write(f"**Presenter:** {row.get('presenter_name') or '-'}")
                st.write(f"**Department:** {row.get('department') or '-'}")
            with c2:
                st.write(f"**Position:** {row.get('position') or '-'}")
                st.write(f"**Training Hours:** {row.get('training_hours') or '-'}")

            st.markdown("**Ratings**")
            for field, label in RATING_QUESTIONS:
                val = row.get(field)
                st.write(f"- {label} → **{val if val is not None else '-'}**/4")

            open_answers = [(label, row.get(field)) for field, label in OPEN_QUESTIONS if row.get(field)]
            if open_answers:
                st.markdown("**Feedback**")
                for label, val in open_answers:
                    st.write(f"- {label}")
                    st.write(val)

            if allow_delete and sb is not None:
                st.divider()
                render_delete_submission_control(sb, row["id"], key_prefix="submlist")


def render_my_submissions_tab(sb: Client):
    """Lets a trainee see the surveys they personally submitted, with dates.
    RLS (responses_select_own) already restricts this query to their own
    rows even without the explicit filter, but we filter anyway for clarity.
    """
    user_id = st.session_state.auth_user["id"]

    if st.button("Refresh", key="refresh_my_submissions"):
        st.rerun()

    try:
        rows = fetch_submissions_for_user(sb, user_id)
    except Exception as e:
        st.error(f"Could not load your submissions: {e}")
        return

    render_submission_list(rows)  # allow_delete stays False: trainees can't delete their own records


# ---------------------------------------------------------------------------
# Admin dashboard
# ---------------------------------------------------------------------------
QUESTION_LABELS = {
    "q1_enjoyable": "Q1 Enjoyable",
    "q2_organized": "Q2 Organized",
    "q3_satisfied": "Q3 Satisfied",
    "q4_improved_understanding": "Q4 Understanding",
    "q5_respect_ideas": "Q5 Respect Ideas",
    "q6_presenter_rating": "Q6 Presenter",
}


@st.cache_data(ttl=60, show_spinner=False)
def fetch_responses(_sb: Client, cache_key: str) -> pd.DataFrame:
    resp = _sb.table("survey_responses").select("*").order("created_at", desc=True).execute()
    df = pd.DataFrame(resp.data)
    if not df.empty:
        # Stored/fetched in UTC; converted to local time once, here, so
        # every display and export downstream (Analytics table, Browse
        # Submissions, Compliance, CSV/Excel) shows Myanmar time without
        # each of them needing to remember to convert it themselves.
        df["created_at"] = pd.to_datetime(df["created_at"]).dt.tz_convert(LOCAL_TZ)
    return df


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    df = df.copy()
    # Excel/openpyxl cannot store timezone-aware datetimes; Supabase returns
    # created_at with a UTC offset, so strip the tz before writing.
    for col in df.columns:
        if isinstance(df[col].dtype, pd.DatetimeTZDtype):
            df[col] = df[col].dt.tz_localize(None)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Responses")
        worksheet = writer.sheets["Responses"]
        worksheet.sheet_view.showGridLines = True

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="Calibri")
        data_alignment = Alignment(vertical="center")
        thin_side = Side(style="thin", color="D9D9D9")
        data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        n_rows, n_cols = df.shape

        worksheet.row_dimensions[1].height = 28
        for col_idx in range(1, n_cols + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row_idx in range(2, n_rows + 2):
            worksheet.row_dimensions[row_idx].height = 22
            for col_idx in range(1, n_cols + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border

        def _cell_len(value) -> int:
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return 0
            return len(str(value))

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max([_cell_len(col_name)] + [_cell_len(v) for v in df.iloc[:, col_idx - 1]])
            width = min(max(max_len + 2, 12), 45)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    return output.getvalue()


def render_admin_dashboard(sb: Client):
    st.markdown(
        f"""
        <div class="survey-header">
            <h1>{COMPANY_NAME}</h1>
            <h3>Admin Analytics Dashboard</h3>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.write(f"Logged in as **{st.session_state.auth_user['email']}** (admin)")
        if st.button("🔄 Refresh", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
        if st.button("Log Out", use_container_width=True):
            sign_out(sb)

    try:
        compliance_df, overdue_count = compute_compliance(sb)
    except Exception:
        compliance_df, overdue_count = None, None  # surfaced again, with the real error, inside the tab

    if overdue_count:
        st.warning(
            f"⚠️ {overdue_count} trainee(s) have gone 7+ days without a survey submission "
            "(or never submitted). See the Compliance tab for details."
        )

    tab_analytics, tab_browse, tab_compliance, tab_users = st.tabs(
        ["📊 Analytics", "📄 Browse Submissions", "⚠️ Compliance", "🗂️ Manage Users"]
    )
    with tab_analytics:
        render_analytics_tab(sb)
    with tab_browse:
        render_browse_submissions_tab(sb)
    with tab_compliance:
        render_compliance_tab(sb, compliance_df)
    with tab_users:
        render_manage_users_tab(sb)


COMPLIANCE_WINDOW_DAYS = 7


def compute_compliance(sb: Client) -> tuple[pd.DataFrame, int]:
    """For every trainee (non-admin user), finds their most recent
    submission and flags them overdue if it's 7+ days old (in LOCAL_TZ) or
    they've never submitted. New accounts get a grace period: an account
    with zero submissions isn't flagged until the account itself is at
    least 7 days old, so a trainee who signed up an hour ago isn't
    immediately shown as non-compliant."""
    users_df = fetch_users(sb, "manage_users")
    responses_df = fetch_responses(sb, st.session_state.auth_user["id"])

    trainees_df = users_df[users_df["role"] != "admin"] if not users_df.empty else users_df
    if trainees_df.empty:
        return pd.DataFrame(), 0

    now_local = datetime.now(LOCAL_TZ)
    cutoff = now_local - timedelta(days=COMPLIANCE_WINDOW_DAYS)

    if not responses_df.empty:
        last_submission = responses_df.groupby("user_id")["created_at"].max()
    else:
        last_submission = pd.Series(dtype="object")

    rows = []
    for _, u in trainees_df.iterrows():
        last = last_submission.get(u["id"])

        if last is not None and pd.notna(last):
            last_local = last.tz_convert(LOCAL_TZ)
            days_since = (now_local - last_local).days
            overdue = last_local < cutoff
            last_str = last_local.strftime("%Y-%m-%d %H:%M")
        else:
            last_str = "Never"
            days_since = None
            account_created = pd.to_datetime(u.get("created_at"), utc=True, errors="coerce")
            overdue = pd.isna(account_created) or account_created.tz_convert(LOCAL_TZ) < cutoff

        rows.append(
            {
                "Name": _s(u.get("full_name")) or "(no name)",
                "Email": _s(u.get("email")),
                "Department": _s(u.get("department")),
                "Last Submission": last_str,
                "_days_since_sort": days_since if days_since is not None else 10**9,
                "Days Since": days_since if days_since is not None else "Never",
                "Overdue": overdue,
            }
        )

    result_df = pd.DataFrame(rows)
    result_df = result_df.sort_values(
        ["Overdue", "_days_since_sort"], ascending=[False, False]
    ).drop(columns="_days_since_sort")
    overdue_count = int(result_df["Overdue"].sum())
    result_df["Overdue"] = result_df["Overdue"].map({True: "⚠️ Yes", False: "✅ No"})
    return result_df, overdue_count


def render_compliance_tab(sb: Client, compliance_df: Optional[pd.DataFrame] = None):
    if compliance_df is None:
        try:
            compliance_df, _ = compute_compliance(sb)
        except Exception as e:
            st.error(f"Could not compute compliance: {e}")
            return

    if compliance_df.empty:
        st.info("No trainee accounts found.")
        return

    overdue_count = (compliance_df["Overdue"] == "⚠️ Yes").sum()
    st.metric(f"Overdue ({COMPLIANCE_WINDOW_DAYS}+ days without a submission)", f"{overdue_count} / {len(compliance_df)}")
    st.caption('Sorted most-overdue first. "Never" means the account has no submissions at all.')
    st.dataframe(compliance_df, use_container_width=True, hide_index=True)


def render_analytics_tab(sb: Client):
    try:
        df = fetch_responses(sb, st.session_state.auth_user["id"])
    except Exception as e:
        st.error(f"Could not load responses: {e}")
        return

    if df.empty:
        st.info("No survey responses yet.")
        return

    # fetch_responses() already converts created_at to LOCAL_TZ, so this is
    # just extracting the calendar date for the Date Range filter to
    # compare against -- otherwise a submission made late at night local
    # time could fall on the wrong side of a range that should include it.
    df["_local_date"] = df["created_at"].dt.date

    # ---- Filters ----
    st.subheader("Filters")
    f1, f2, f3 = st.columns(3)
    with f1:
        courses = sorted(df["training_course"].dropna().unique().tolist())
        selected_courses = st.multiselect("Training Course", courses)
    with f2:
        departments = sorted(df["department"].dropna().unique().tolist())
        selected_departments = st.multiselect("Department", departments)
    with f3:
        min_date = df["_local_date"].min()
        max_date = df["_local_date"].max()
        date_range = st.date_input(
            "Date Range", value=(min_date, max_date), min_value=min_date, max_value=max_date
        )

    filtered = df.copy()
    if selected_courses:
        filtered = filtered[filtered["training_course"].isin(selected_courses)]
    if selected_departments:
        filtered = filtered[filtered["department"].isin(selected_departments)]
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_d, end_d = date_range
        filtered = filtered[(filtered["_local_date"] >= start_d) & (filtered["_local_date"] <= end_d)]

    if filtered.empty:
        st.warning("No responses match the selected filters.")
        return

    rating_cols = list(QUESTION_LABELS.keys())
    filtered["overall_avg"] = filtered[rating_cols].mean(axis=1)

    # ---- Overview metrics ----
    st.subheader("Overview")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Submissions", len(filtered))
    m2.metric("Avg Overall Rating", f"{filtered['overall_avg'].mean():.2f} / 4")
    m3.metric("Courses Covered", filtered["training_course"].nunique())

    top_presenter_df = (
        filtered.dropna(subset=["presenter_name"])
        .groupby("presenter_name")["q6_presenter_rating"]
        .mean()
        .sort_values(ascending=False)
    )
    top_presenter = top_presenter_df.index[0] if not top_presenter_df.empty else "N/A"
    m4.metric("Top Rated Presenter", top_presenter)

    # ---- Average rating per course ----
    st.subheader("Average Rating by Course")
    course_avg = (
        filtered.groupby("training_course")["overall_avg"].mean().sort_values(ascending=False).round(2)
    )
    st.bar_chart(course_avg, color=EMERALD)

    # ---- Criteria ratings Q1-Q6 ----
    st.subheader("Average Rating by Criteria (Q1–Q6)")
    criteria_avg = filtered[rating_cols].mean().round(2)
    criteria_avg.index = [QUESTION_LABELS[c] for c in criteria_avg.index]
    st.bar_chart(criteria_avg, color=SLATE)

    # ---- Top presenters table ----
    st.subheader("Top Rated Presenters")
    st.dataframe(
        top_presenter_df.round(2).reset_index().rename(
            columns={"presenter_name": "Presenter", "q6_presenter_rating": "Avg Rating"}
        ),
        use_container_width=True,
        hide_index=True,
    )

    # ---- Data table ----
    st.subheader("All Submissions")
    display_cols = [
        "created_at", "trainee_name", "department", "training_course", "presenter_name",
        "overall_avg",
    ] + rating_cols + [f for f, _ in OPEN_QUESTIONS]
    st.dataframe(filtered[display_cols], use_container_width=True, hide_index=True)

    # ---- Export ----
    st.subheader("Export")
    e1, e2 = st.columns(2)
    with e1:
        st.download_button(
            "Download CSV",
            data=filtered[display_cols].to_csv(index=False).encode("utf-8"),
            file_name=f"survey_responses_{datetime.now():%Y%m%d_%H%M%S}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with e2:
        st.download_button(
            "Download Excel (.xlsx)",
            data=to_excel_bytes(filtered[display_cols]),
            file_name=f"survey_responses_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ---------------------------------------------------------------------------
# Admin: browse submissions one-by-one (mirrors the trainee's own form view)
# ---------------------------------------------------------------------------
RATING_LABELS = {4: "Excellent", 3: "Good", 2: "Average", 1: "Poor"}


def render_browse_submissions_tab(sb: Client):
    try:
        df = fetch_responses(sb, st.session_state.auth_user["id"])
    except Exception as e:
        st.error(f"Could not load responses: {e}")
        return

    if df.empty:
        st.info("No survey responses yet.")
        return

    df = df.sort_values("created_at", ascending=False).reset_index(drop=True)

    # ---- Filter by the submitter's actual account (not the free-text
    # "Trainee Name" field, which a trainee could type differently across
    # submissions), so an admin reliably pages through one person's forms.
    try:
        users_df = fetch_users(sb, "manage_users")
    except Exception:
        users_df = pd.DataFrame(columns=["id", "email", "full_name"])

    if not users_df.empty:
        df = df.merge(
            users_df[["id", "email", "full_name"]].rename(
                columns={"email": "account_email", "full_name": "account_name"}
            ),
            left_on="user_id",
            right_on="id",
            how="left",
            suffixes=("", "_profile"),
        )
    else:
        df["account_email"] = None
        df["account_name"] = None

    def _account_key(r) -> str:
        email = _s(r.get("account_email"))
        return email if email else f"unlinked:{_s(r.get('trainee_name'))}"

    def _account_label(r) -> str:
        name = _s(r.get("account_name")) or _s(r.get("trainee_name")) or "(no name)"
        email = _s(r.get("account_email"))
        return f"{name} ({email})" if email else f"{name} — no linked account"

    df["_account_key"] = df.apply(_account_key, axis=1)
    df["_account_label"] = df.apply(_account_label, axis=1)

    accounts = df[["_account_key", "_account_label"]].drop_duplicates().sort_values("_account_label")
    account_options = ["All Trainees"] + accounts["_account_key"].tolist()
    account_label_map = {"All Trainees": "All Trainees", **dict(zip(accounts["_account_key"], accounts["_account_label"]))}

    selected_account = st.selectbox(
        "View submissions from",
        account_options,
        format_func=lambda k: account_label_map.get(k, k),
    )

    view_df = (
        df[df["_account_key"] == selected_account].reset_index(drop=True)
        if selected_account != "All Trainees"
        else df
    )
    n = len(view_df)
    if n == 0:
        st.info("No submissions for this account.")
        return

    # Reset to the first submission whenever the account filter changes.
    if st.session_state.get("browse_account_prev") != selected_account:
        st.session_state.browse_idx = 0
        st.session_state.browse_account_prev = selected_account
    elif "browse_idx" not in st.session_state or st.session_state.browse_idx >= n:
        st.session_state.browse_idx = 0

    def _label(i: int) -> str:
        r = view_df.iloc[i]
        created = r["created_at"].strftime("%Y-%m-%d %H:%M")
        return f"{i + 1}/{n} — {_s(r.get('trainee_name')) or '(no name)'} — {created}"

    # Evaluate both nav buttons (and any resulting session_state change +
    # rerun) before the selectbox below is instantiated with the same key.
    nav1, nav2, nav3 = st.columns([1, 3, 1])
    prev_clicked = nav1.button(
        "⬅ Prev", use_container_width=True, disabled=st.session_state.browse_idx == 0
    )
    next_clicked = nav3.button(
        "Next ➡", use_container_width=True, disabled=st.session_state.browse_idx >= n - 1
    )
    if prev_clicked:
        st.session_state.browse_idx -= 1
        st.rerun()
    if next_clicked:
        st.session_state.browse_idx += 1
        st.rerun()

    with nav2:
        st.selectbox(
            "Jump to submission",
            options=list(range(n)),
            format_func=_label,
            key="browse_idx",
            label_visibility="collapsed",
        )

    st.divider()
    row = view_df.iloc[st.session_state.browse_idx]

    st.markdown(f"### {_s(row.get('trainee_name')) or '(no name)'}")
    account_email = _s(row.get("account_email"))
    submitted_at = row["created_at"].strftime("%Y-%m-%d %H:%M")
    st.caption(f"{account_email + ' · ' if account_email else ''}Submitted {submitted_at}")

    c1, c2 = st.columns(2)
    with c1:
        st.write(f"**Position:** {_s(row.get('position')) or '-'}")
        st.write(f"**Department:** {_s(row.get('department')) or '-'}")
        st.write(f"**Training Course:** {_s(row.get('training_course')) or '-'}")
        st.write(f"**Presenter:** {_s(row.get('presenter_name')) or '-'}")
    with c2:
        st.write(f"**Start Date:** {_s(row.get('start_date')) or '-'}")
        st.write(f"**Completed Date:** {_s(row.get('completed_date')) or '-'}")
        st.write(f"**Training Hours:** {_s(row.get('training_hours')) or '-'}")
        st.write(f"**Time Range:** {_s(row.get('time_range')) or '-'}")

    st.divider()
    st.markdown("#### Ratings")
    for field, label in RATING_QUESTIONS:
        val = row.get(field)
        if pd.notna(val):
            val_int = int(val)
            text = f"{val_int} - {RATING_LABELS.get(val_int, '')}"
        else:
            text = "-"
        st.write(f"- {label} → **{text}**")

    st.divider()
    st.markdown("#### Feedback")
    for field, label in OPEN_QUESTIONS:
        val = _s(row.get(field))
        st.markdown(f"**{label}**")
        st.write(val if val else "_No response_")

    st.divider()
    render_delete_submission_control(sb, row["id"], key_prefix="browse")


# ---------------------------------------------------------------------------
# Admin: manage users
# ---------------------------------------------------------------------------
@st.cache_data(ttl=30, show_spinner=False)
def fetch_users(_sb: Client, cache_key: str) -> pd.DataFrame:
    resp = _sb.table("users_profile").select("*").order("email").execute()
    return pd.DataFrame(resp.data)


def _s(value) -> str:
    """Safely stringify a pandas row value, treating NaN/None as empty.
    pandas represents missing strings as float NaN, not None, so a plain
    `value or default` check isn't enough — NaN is truthy in Python."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value)


def render_manage_users_tab(sb: Client):
    """Lets an admin view every user and promote/demote their role.
    Backed by the profile_update_admin RLS policy, which lets any admin
    update any users_profile row using the regular publishable key.
    """
    if st.button("Refresh Users", key="refresh_users"):
        st.cache_data.clear()
        st.rerun()

    try:
        users_df = fetch_users(sb, "manage_users")
    except Exception as e:
        st.error(f"Could not load users: {e}")
        return

    if users_df.empty:
        st.info("No users found.")
        return

    st.caption(f"{len(users_df)} user(s)")

    # Detect this up front rather than letting every Delete button fail
    # silently one at a time — this is the most common reason deletion
    # doesn't work, so surface it once, clearly, at the top of the tab.
    delete_enabled = bool(st.secrets.get("supabase", {}).get("service_role_key"))
    if not delete_enabled:
        st.info(
            "User deletion is turned off: no `service_role_key` is configured in secrets. "
            "See README.md section 4 to enable it."
        )
    current_admin_id = st.session_state.auth_user["id"]

    for _, row in users_df.iterrows():
        uid = row["id"]
        current_role = _s(row.get("role")) or "user"
        is_self = uid == current_admin_id

        with st.container(border=True):
            c1, c2, c3 = st.columns([3, 2, 2])
            with c1:
                st.write(f"**{_s(row.get('full_name')) or '(no name)'}**")
                st.caption(_s(row.get("email")))
                extra = " · ".join(p for p in (_s(row.get("department")), _s(row.get("position"))) if p)
                if extra:
                    st.caption(extra)
            with c2:
                new_role = st.segmented_control(
                    "Role",
                    options=["user", "admin"],
                    default=current_role,
                    key=f"role_{uid}",
                    label_visibility="collapsed",
                    disabled=is_self,
                )
            with c3:
                if is_self:
                    st.caption("This is you")
                elif st.button("Save", key=f"save_{uid}", use_container_width=True):
                    if new_role == current_role:
                        st.info("No change.")
                    else:
                        try:
                            sb.table("users_profile").update({"role": new_role}).eq("id", uid).execute()
                            st.cache_data.clear()
                            st.toast(f"{_s(row.get('email'))} is now {new_role}", icon="✅")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Could not update role: {e}")

            email = _s(row.get("email"))
            is_viewing = st.session_state.get("manage_viewing_uid") == uid
            is_editing = st.session_state.get("manage_editing_uid") == uid
            is_confirming_delete = st.session_state.get("manage_confirm_delete_uid") == uid

            b1, b2, b3 = st.columns(3)
            with b1:
                if st.button(
                    "🗂️ Hide Submissions" if is_viewing else "🗂️ View Submissions",
                    key=f"view_{uid}",
                    use_container_width=True,
                ):
                    st.session_state.manage_viewing_uid = None if is_viewing else uid
                    st.rerun()
            with b2:
                if st.button(
                    "✏️ Cancel Edit" if is_editing else "✏️ Edit Info",
                    key=f"edit_{uid}",
                    use_container_width=True,
                ):
                    st.session_state.manage_editing_uid = None if is_editing else uid
                    st.rerun()
            with b3:
                if is_self:
                    st.caption("Can't delete your own account")
                elif not delete_enabled:
                    st.caption("Deletion disabled (see note above)")
                elif st.button("🗑️ Delete User", key=f"delete_{uid}", use_container_width=True):
                    st.session_state.manage_confirm_delete_uid = uid
                    st.rerun()

            if is_editing:
                st.divider()
                st.caption("Fixes a blank profile (e.g. \"(no name)\") for accounts that never got their info saved at signup.")
                new_name = st.text_input("Full Name", value=_s(row.get("full_name")), key=f"edit_name_{uid}")
                ec1, ec2 = st.columns(2)
                with ec1:
                    new_dept = st.text_input(
                        "Department", value=_s(row.get("department")), key=f"edit_dept_{uid}"
                    )
                with ec2:
                    new_pos = st.text_input(
                        "Position", value=_s(row.get("position")), key=f"edit_pos_{uid}"
                    )
                if st.button("Save Profile", key=f"save_profile_{uid}", type="primary"):
                    try:
                        sb.table("users_profile").update(
                            {"full_name": new_name.strip(), "department": new_dept.strip(), "position": new_pos.strip()}
                        ).eq("id", uid).execute()
                        st.cache_data.clear()
                        st.session_state.manage_editing_uid = None
                        if is_self:
                            st.session_state.profile = load_profile(sb, uid)
                        st.toast(f"Updated {email}'s profile", icon="✅")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not update profile ({type(e).__name__}): {e}")

            if is_viewing:
                st.divider()
                st.markdown(f"**Submissions from {email}** — viewed exactly as {email} sees their own")
                try:
                    submissions = fetch_submissions_for_user(sb, uid)
                except Exception as e:
                    st.error(f"Could not load submissions: {e}")
                else:
                    render_submission_list(submissions, sb=sb, allow_delete=True)

            if is_confirming_delete:
                st.divider()
                st.warning(
                    f"Permanently delete **{email}**? This cannot be undone. "
                    "Their past survey submissions are kept but unlinked from any account."
                )
                d1, d2 = st.columns(2)
                with d1:
                    confirm_clicked = st.button(
                        "Yes, delete permanently",
                        key=f"confirm_delete_{uid}",
                        use_container_width=True,
                        type="primary",
                    )
                with d2:
                    if st.button("Cancel", key=f"cancel_delete_{uid}", use_container_width=True):
                        st.session_state.manage_confirm_delete_uid = None
                        st.rerun()

                # Rendered full-width below the columns (not squeezed inside
                # d1) so a failure is impossible to miss.
                if confirm_clicked:
                    try:
                        delete_user(uid)
                        st.cache_data.clear()
                        st.session_state.manage_confirm_delete_uid = None
                        st.toast(f"Deleted {email}", icon="🗑️")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not delete user ({type(e).__name__}): {e}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    init_session_state()
    sb = get_supabase()

    # The app runs in wide layout so the admin dashboard's tables/charts get
    # full desktop width; trainee-facing pages opt into a centered, narrower
    # "card" width via this keyed container (styled as .st-key-app-shell in
    # CUSTOM_CSS) instead of constraining the whole app.
    if st.session_state.auth_user is None:
        with st.container(key="app-shell"):
            render_auth_screen(sb)
        return

    if st.session_state.profile is None:
        try:
            st.session_state.profile = load_profile(sb, st.session_state.auth_user["id"])
        except Exception as e:
            with st.container(key="app-shell"):
                st.error(f"Could not load your profile: {e}")
                if st.button("Log Out"):
                    sign_out(sb)
            return

    role = (st.session_state.profile or {}).get("role", "user")
    if role == "admin":
        render_admin_dashboard(sb)
    else:
        with st.container(key="app-shell"):
            render_survey_form(sb)


if __name__ == "__main__":
    main()
